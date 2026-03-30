#!/usr/bin/env python3
"""
特定保険医療材料のレセプトコメント一覧（別表Ⅱ）を生成するスクリプト。
入力: 原文/別表Ⅳ 診療報酬明細書の「摘要」欄への記載事項等一覧（材料価格基準）.xlsx
出力: 特定保険医療材料_レセプトコメント一覧（別表Ⅱ）.xlsx
"""

import openpyxl
import xlsxwriter
import sys

INPUT_FILE = '原文/別表Ⅳ 診療報酬明細書の「摘要」欄への記載事項等一覧（材料価格基準）.xlsx'
OUTPUT_FILE = '特定保険医療材料_レセプトコメント一覧（別表Ⅱ）.xlsx'


def extract_betsu2_data(input_file):
    """別表Ⅱのデータを抽出し、親項目＋子コード行として構造化する。"""
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb[wb.sheetnames[0]]

    items = []
    current_parent = None
    in_betsu2 = False

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        betsu = str(row[1].value).strip() if row[1].value else ''
        item_no = str(row[0].value).strip() if row[0].value else ''
        kubun = str(row[2].value).strip() if row[2].value else ''
        kinou = str(row[3].value).strip() if row[3].value else ''
        kisai = str(row[4].value).strip() if row[4].value else ''
        code = str(row[5].value).strip() if row[5].value else ''
        hyouji = str(row[6].value).strip() if row[6].value else ''

        # 別表Ⅱの開始・終了を追跡
        if betsu == 'Ⅱ':
            in_betsu2 = True
        elif betsu and betsu != 'Ⅱ':
            in_betsu2 = False
            current_parent = None
            continue

        if not in_betsu2:
            continue

        if item_no:
            # 新しい親項目
            current_parent = {
                'item_no': item_no,
                'kubun': kubun,
                'kinou': kinou,
                'kisai': kisai,
                'rows': [(code, hyouji)],
            }
            items.append(current_parent)
        elif current_parent is not None:
            # 子行
            current_parent['rows'].append((code, hyouji))
            if kisai:
                current_parent['kisai'] += '\n' + kisai

    wb.close()
    return items


def classify_code(code):
    """コード先頭3桁でコメント種別を判定する。"""
    if code.startswith('83'):
        return 'フリーコメント'
    elif code.startswith('82'):
        return '選択式コメント'
    elif code.startswith('85'):
        return '日付記載'
    elif code.startswith('84'):
        return '数値記載'
    else:
        return 'その他'


def write_excel(items, output_file):
    """一覧表と運用変更点まとめの2シートをExcelに出力する。"""
    wb = xlsxwriter.Workbook(output_file)

    # ── 書式定義 ──
    title_fmt = wb.add_format({
        'bold': True, 'font_size': 14, 'font_name': '游ゴシック',
        'align': 'left', 'valign': 'vcenter',
    })
    header_fmt = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': '游ゴシック',
        'bg_color': '#4472C4', 'font_color': 'white',
        'border': 1, 'text_wrap': True,
        'align': 'center', 'valign': 'vcenter',
    })
    cell_fmt = wb.add_format({
        'font_size': 10, 'font_name': '游ゴシック',
        'border': 1, 'text_wrap': True, 'valign': 'top',
    })
    center_fmt = wb.add_format({
        'font_size': 10, 'font_name': '游ゴシック',
        'border': 1, 'text_wrap': True,
        'align': 'center', 'valign': 'top',
    })
    new_fmt = wb.add_format({
        'font_size': 10, 'font_name': '游ゴシック',
        'border': 1, 'text_wrap': True,
        'align': 'center', 'valign': 'top',
        'bg_color': '#FFF2CC',  # 薄黄色で新設を強調
    })
    # 運用まとめ用
    summary_header_fmt = wb.add_format({
        'bold': True, 'font_size': 10, 'font_name': '游ゴシック',
        'bg_color': '#548235', 'font_color': 'white',
        'border': 1, 'text_wrap': True,
        'align': 'center', 'valign': 'vcenter',
    })
    summary_cell_fmt = wb.add_format({
        'font_size': 10, 'font_name': '游ゴシック',
        'border': 1, 'text_wrap': True, 'valign': 'top',
    })
    summary_center_fmt = wb.add_format({
        'font_size': 10, 'font_name': '游ゴシック',
        'border': 1, 'text_wrap': True,
        'align': 'center', 'valign': 'top',
    })
    attention_fmt = wb.add_format({
        'font_size': 10, 'font_name': '游ゴシック',
        'border': 1, 'text_wrap': True, 'valign': 'top',
        'bg_color': '#FCE4EC',  # 薄赤 = 注意
    })

    # ========================================
    # シート1: レセプトコメント一覧
    # ========================================
    ws1 = wb.add_worksheet('レセプトコメント一覧')

    # タイトル
    ws1.merge_range('A1:I1',
                    '【令和８年度新設】特定保険医療材料 レセプトコメント一覧（別表Ⅱ）',
                    title_fmt)
    ws1.set_row(0, 25)

    # 注記
    note_fmt = wb.add_format({
        'font_size': 9, 'font_name': '游ゴシック',
        'italic': True, 'font_color': '#C00000',
    })
    ws1.merge_range('A2:I2',
                    '※ 令和８年６月１日適用。以下すべて新設項目。算定時にレセプト摘要欄への記載が必須となります。',
                    note_fmt)

    # ヘッダー
    headers = ['項番', '区分', '特定保険医療材料の機能区分', '記載事項',
               'レセプト電算コード', 'レセプト表示文言', 'コメント種別',
               '適用開始', '備考']
    col_widths = [6, 6, 35, 40, 14, 50, 14, 10, 15]
    for ci, (h, w) in enumerate(zip(headers, col_widths)):
        ws1.write(2, ci, h, header_fmt)
        ws1.set_column(ci, ci, w)

    # データ行
    row_idx = 3
    for item in items:
        n_rows = len(item['rows'])
        for ri, (code, hyouji) in enumerate(item['rows']):
            if ri == 0:
                # 親行の項番・区分・機能区分・記載事項を書く
                if n_rows > 1:
                    ws1.merge_range(row_idx, 0, row_idx + n_rows - 1, 0,
                                    item['item_no'], center_fmt)
                    ws1.merge_range(row_idx, 1, row_idx + n_rows - 1, 1,
                                    item['kubun'], center_fmt)
                    ws1.merge_range(row_idx, 2, row_idx + n_rows - 1, 2,
                                    item['kinou'], cell_fmt)
                    ws1.merge_range(row_idx, 3, row_idx + n_rows - 1, 3,
                                    item['kisai'], cell_fmt)
                    ws1.merge_range(row_idx, 7, row_idx + n_rows - 1, 7,
                                    'R8.6.1', new_fmt)
                    ws1.merge_range(row_idx, 8, row_idx + n_rows - 1, 8,
                                    '新設', new_fmt)
                else:
                    ws1.write(row_idx, 0, item['item_no'], center_fmt)
                    ws1.write(row_idx, 1, item['kubun'], center_fmt)
                    ws1.write(row_idx, 2, item['kinou'], cell_fmt)
                    ws1.write(row_idx, 3, item['kisai'], cell_fmt)
                    ws1.write(row_idx, 7, 'R8.6.1', new_fmt)
                    ws1.write(row_idx, 8, '新設', new_fmt)

            ws1.write(row_idx + ri, 4, code, center_fmt)
            ws1.write(row_idx + ri, 5, hyouji, cell_fmt)
            ws1.write(row_idx + ri, 6, classify_code(code), center_fmt)

        row_idx += n_rows

    # ========================================
    # シート2: 運用変更点まとめ
    # ========================================
    ws2 = wb.add_worksheet('運用変更点まとめ')

    ws2.merge_range('A1:F1',
                    '【医事課向け】特定保険医療材料 レセプトコメント運用変更点（R8新設）',
                    title_fmt)
    ws2.set_row(0, 25)

    ws2.merge_range('A2:F2',
                    '※ 令和８年６月１日より、以下の特定保険医療材料を算定する際に新たにレセプト摘要欄への記載が必要になります。',
                    note_fmt)

    # ヘッダー
    s_headers = ['区分', '機能区分名', 'コメント数', 'コメント種別',
                 '主な記載内容', '注意点']
    s_widths = [6, 35, 10, 18, 50, 35]
    for ci, (h, w) in enumerate(zip(s_headers, s_widths)):
        ws2.write(2, ci, h, summary_header_fmt)
        ws2.set_column(ci, ci, w)

    row_idx = 3
    for item in items:
        codes = item['rows']
        types = set(classify_code(c) for c, _ in codes)
        type_str = '／'.join(sorted(types))

        # 主な記載内容を要約（記載事項の先頭部分）
        kisai_summary = item['kisai'].replace('\n', ' ').strip()
        if len(kisai_summary) > 100:
            kisai_summary = kisai_summary[:100] + '…'

        # 注意点を判定
        notes = []
        if any('日' in h or '年月' in h for _, h in codes):
            notes.append('日付入力あり')
        if any('詳記' in h or '詳細' in h for _, h in codes):
            notes.append('症状詳記が必要')
        if any('選択' in item['kisai'] or 'いずれ' in item['kisai']
               for _ in [1]):
            notes.append('選択肢あり')
        if len(codes) >= 3:
            notes.append(f'コード{len(codes)}件（複数入力）')

        # 注意点が多い項目は赤背景
        fmt = attention_fmt if len(notes) >= 2 else summary_cell_fmt

        ws2.write(row_idx, 0, item['kubun'], summary_center_fmt)
        ws2.write(row_idx, 1, item['kinou'], fmt)
        ws2.write(row_idx, 2, len(codes), summary_center_fmt)
        ws2.write(row_idx, 3, type_str, summary_center_fmt)
        ws2.write(row_idx, 4, kisai_summary, fmt)
        ws2.write(row_idx, 5, '\n'.join(notes) if notes else '－', fmt)
        row_idx += 1

    # ========================================
    # シート3: コード種別凡例
    # ========================================
    ws3 = wb.add_worksheet('凡例')
    ws3.write(0, 0, 'コード種別の見方', title_fmt)
    legend = [
        ('コード先頭', '種別', '説明'),
        ('83xxxxx', 'フリーコメント', '自由記載（文字列入力）が必要'),
        ('82xxxxx', '選択式コメント', '所定の選択肢から該当するものを選択'),
        ('85xxxxx', '日付記載', '年月日の入力が必要'),
        ('84xxxxx', '数値記載', '数値の入力が必要'),
    ]
    for ri, (a, b, c) in enumerate(legend):
        fmt = header_fmt if ri == 0 else cell_fmt
        ws3.write(ri + 1, 0, a, fmt)
        ws3.write(ri + 1, 1, b, fmt)
        ws3.write(ri + 1, 2, c, fmt)
    ws3.set_column(0, 0, 14)
    ws3.set_column(1, 1, 18)
    ws3.set_column(2, 2, 50)

    wb.close()
    return row_idx - 3  # 運用まとめの行数


def main():
    print('別表Ⅱデータを抽出中...')
    items = extract_betsu2_data(INPUT_FILE)
    print(f'  親項目数: {len(items)}')
    print(f'  総コード数: {sum(len(i["rows"]) for i in items)}')

    print(f'Excel出力中: {OUTPUT_FILE}')
    summary_count = write_excel(items, OUTPUT_FILE)
    print(f'  シート1: レセプトコメント一覧')
    print(f'  シート2: 運用変更点まとめ ({summary_count}件)')
    print(f'  シート3: 凡例')
    print('完了!')


if __name__ == '__main__':
    main()
